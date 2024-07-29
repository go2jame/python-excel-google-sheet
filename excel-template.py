#excel-template.py

from openpyxl import load_workbook

data_excel = load_workbook(filename='data-invoice.xlsx')
sheet_data = data_excel.active  # data_excel['ใบกำกับภาษี']

# print(sheet_data['E2'].value)

# number = sheet_data['E2'].value
# # แก้ไข cell ใน template
# sheet_template['S14'].value = number
# template_excel.save('resault-excel.xlsx')

# ------
count  = len(sheet_data['A'])
# print(count)

rows = []

for i in range(2,count+1):
    datalist = []
    for d in sheet_data[i]:
        datalist.append(d.value)

    rows.append(datalist)

# loop เพื่อสร้างกลุ่ม IV
rows_dict = {}
for r in rows:
    if r[4] not in rows_dict:
        rows_dict[r[4]] = []

# loop เพื่อแยก IV แต่ละ transacetion
for r in rows:
    rows_dict[r[4]].append(r)

##############
count_name = 1

for rw in rows_dict.values():
    template_excel = load_workbook(filename='template-iv.xlsx')
    sheet_template = template_excel.active # template_excel[0]
    # rw =  [['บริษัท A,....],['บริษัท A,....],['บริษัท A,....]]
    row_product = 25
    
    check_company = False
    check_address1 = False
    check_address2 = False
    check_taxid = False
    check_number = False
    check_date = False

    for i,r in enumerate(rw,start=1):
        data = r    #rows_dist['IV-6307021'][0]
        company = data[0]
        address1 = data[1]
        address2 = data[2]
        taxid = data[3]
        number = data[4]
        date = data[9]

        if check_company == False:
            sheet_template['C13'] = company
            check_company = True
        if check_address1 == False:
            sheet_template['C15'] = address1
            check_address1 = True
        if check_address2 == False:
            sheet_template['C17'] = address2
            check_address2 = True
        if check_taxid == False:
            sheet_template['D20'] = taxid
            check_taxid = True
        if check_number == False:
            sheet_template['S14'] = number
            check_number = True
        if check_date == False:
            sheet_template['S16'] = date
            check_date = True

        product_no = data[5]
        product_item = data[6]
        product_quantity = data[7]
        product_price= data[8]

        sheet_template.cell(row=row_product,column=2).value = i
        sheet_template.cell(row=row_product,column=3).value = product_no
        sheet_template.cell(row=row_product,column=5).value = product_item
        sheet_template.cell(row=row_product,column=11).value = product_quantity
        sheet_template.cell(row=row_product,column=14).value = 'PCS'
        sheet_template.cell(row=row_product,column=17).value = product_price
        row_product = row_product + 1

        # record 2
        # data = rows_dict['IV-6307021'][1]
        # product_no = data[5]
        # product_item = data[6]
        # product_quantity = data[7]
        # product_price= data[8]
        # sheet_template.cell(row=26,column=3).value = product_no
        # sheet_template.cell(row=26,column=5).value = product_item
        # sheet_template.cell(row=26,column=11).value = product_quantity
        # sheet_template.cell(row=26,column=14).value = 'PCS'
        # sheet_template.cell(row=26,column=17).value = product_price

    template_excel.save('result-data-template-{}.xlsx'.format(number))
    count_name += 1
